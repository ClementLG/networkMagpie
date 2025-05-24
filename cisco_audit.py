#!/usr/bin/env python3

import json
import csv
import os
import datetime
import re
from netmiko import ConnectHandler
from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException, SSHException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import traceback

# --- Configuration des couleurs pour Excel ---
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BOLD_FONT = Font(bold=True)


# --- Fonctions d'aide pour Excel ---
def apply_header_style(ws, row_num=1):
    for cell in ws[row_num]:
        cell.font = BOLD_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


def set_cell_status_color(cell, level):
    if level == "good":
        cell.fill = GREEN_FILL
    elif level == "warning":
        cell.fill = ORANGE_FILL
    elif level == "error" or level == "bad":
        cell.fill = RED_FILL


# --- Fonctions de collecte d'informations ---
def get_device_info(net_connect):
    info = {}
    try:
        output_raw = net_connect.send_command("show version", use_textfsm=True)
        output = output_raw if isinstance(output_raw, list) and output_raw else [{}]
        if output:
            dev_info = output[0]
            info['hostname'] = dev_info.get('hostname', 'N/A')
            info['ios_version'] = dev_info.get('version', 'N/A')
            hardware_info = dev_info.get('hardware', ['N/A'])
            info['model'] = hardware_info[0] if isinstance(hardware_info, list) and hardware_info else 'N/A'
            info['uptime'] = dev_info.get('uptime', 'N/A')
            serial_info = dev_info.get('serial', ['N/A'])
            info['serial_number'] = serial_info[0] if isinstance(serial_info, list) and serial_info else 'N/A'
        if info.get('hostname', 'N/A') == 'N/A' or not info.get('hostname'):
            prompt = net_connect.base_prompt
            if prompt: info['hostname'] = prompt.strip("#> ")
        return info
    except Exception as e:
        print(f"  Erreur get_device_info pour {net_connect.host}: {e}")
        return {'hostname': 'Error', 'ios_version': 'Error', 'model': 'Error', 'uptime': 'Error',
                'serial_number': 'Error'}


def get_interfaces(net_connect):
    interfaces = []
    try:
        ip_interfaces_raw = net_connect.send_command("show ip interface brief", use_textfsm=True)
        ip_interfaces_list = ip_interfaces_raw if isinstance(ip_interfaces_raw, list) else []

        if not ip_interfaces_list:
            # print(f"  [AVERTISSEMENT] Pour {net_connect.host}, 'show ip interface brief' avec TextFSM n'a retourné aucune donnée. Tentative de parsing Regex...")
            raw_output_ip_brief = net_connect.send_command("show ip interface brief", use_textfsm=False)
            ip_interfaces_list = []
            interface_regex = re.compile(
                r"^(?P<interface>\S+)\s+"
                r"(?P<ip_address>[\d\.]+|unassigned)\s+"
                r"\S+\s+\S+\s+"
                r"(?P<status>administratively down|up|down|[^ ]+)"
                r"\s+(?P<proto>up|down)$"
            )
            for line in raw_output_ip_brief.splitlines():
                line = line.strip()
                if not line or line.lower().startswith("interface"): continue
                match = interface_regex.match(line)
                if match:
                    data = match.groupdict()
                    if data['status'].lower() == "administratively" and data['proto'].lower() == "down":
                        data['status'] = "administratively down"
                    ip_interfaces_list.append(data)
            if ip_interfaces_list: print(
                f"  [INFO] Parsing Regex pour 'show ip interface brief' a récupéré {len(ip_interfaces_list)} interfaces.")

        interface_status_raw = net_connect.send_command("show interfaces status", use_textfsm=True)
        interface_descriptions_raw = net_connect.send_command("show interfaces description", use_textfsm=True)
        interface_status_list = interface_status_raw if isinstance(interface_status_raw, list) else []
        interface_descriptions_list = interface_descriptions_raw if isinstance(interface_descriptions_raw, list) else []

        status_map = {item.get('port'): item for item in interface_status_list if item.get('port')}
        desc_map = {item.get('port'): str(item.get('descrip', item.get('description')))
                    for item in interface_descriptions_list
                    if item.get('port') and item.get('descrip', item.get('description')) and
                    str(item.get('descrip', item.get('description'))).strip() and
                    str(item.get('descrip', item.get('description'))) != '--'}

        for iface_data in ip_interfaces_list:
            name = iface_data.get('interface')
            if not name: continue

            protocol_status = iface_data.get('protocol', iface_data.get('proto', 'N/A')).lower()
            link_status_brief = iface_data.get('status', 'N/A').lower()
            ip_address_val = iface_data.get('ip_address', iface_data.get('ipaddr', 'unassigned'))

            phys_status_detail = status_map.get(name, {})
            current_link_status = link_status_brief

            if phys_status_detail:
                phys_status_val = phys_status_detail.get('status', '').lower()
                if phys_status_val:
                    if phys_status_val in ["connected", "disabled", "notconnect", "inactive",
                                           "monitoring"] or "err-disabled" in phys_status_val:
                        current_link_status = phys_status_val
                    if phys_status_val == "disabled": current_link_status = "administratively down"

            if "admin" in link_status_brief and "down" in link_status_brief:
                current_link_status = "administratively down"

            interface_type = "Virtual"
            name_lower = name.lower()
            physical_prefixes = ("eth", "gi", "fa", "te", "twe", "hu", "fo", "se")
            virtual_prefixes = ("vl", "lo", "tu", "po", "bv", "nu", "oo", "gr")
            if any(name_lower.startswith(p) for p in physical_prefixes):
                interface_type = "Physical"
            elif any(name_lower.startswith(p) for p in virtual_prefixes):
                interface_type = "Virtual"

            interfaces.append({
                "name": name, "ip_address": ip_address_val if ip_address_val != 'unassigned' else "N/A",
                "status_link": current_link_status, "status_protocol": protocol_status,
                "description": desc_map.get(name, "N/A"), "type": interface_type,
                "vlan": phys_status_detail.get('vlan', 'N/A'), "duplex": phys_status_detail.get('duplex', 'N/A'),
                "speed": phys_status_detail.get('speed', 'N/A'),
            })
        return interfaces
    except Exception as e:
        print(f"  Erreur critique dans get_interfaces pour {net_connect.host}: {e}")
        traceback.print_exc()
        return []


def get_vlans(net_connect):
    vlans = []
    try:
        output_raw = net_connect.send_command("show vlan brief", use_textfsm=True)
        output = output_raw if isinstance(output_raw, list) else []
        if output:
            for vlan_entry in output:
                ports_list = vlan_entry.get('interfaces', [])
                ports_str = ", ".join(ports_list) if isinstance(ports_list, list) and ports_list else "N/A"
                vlan_name = vlan_entry.get('name', vlan_entry.get('vlan_name', 'N/A'))
                vlans.append({"id": vlan_entry.get('vlan_id', 'N/A'), "name": vlan_name,
                              "status": vlan_entry.get('status', 'N/A'), "ports": ports_str})
        return vlans
    except Exception as e:
        print(f"  Erreur critique dans get_vlans pour {net_connect.host}: {e}"); traceback.print_exc(); return []


def get_arp_table(net_connect):
    arp_table = []
    try:
        output_raw = net_connect.send_command("show ip arp", use_textfsm=True)
        output = output_raw if isinstance(output_raw, list) else []
        if output:
            for entry in output:
                arp_table.append({
                    "protocol": entry.get('protocol', 'N/A'),
                    "address": entry.get('address', entry.get('ip_address', 'N/A')),
                    "age": entry.get('age', 'N/A'),
                    "mac_address": entry.get('mac', entry.get('mac_address', 'N/A')),
                    "type": entry.get('type', 'N/A'), "interface": entry.get('interface', 'N/A')})
        return arp_table
    except Exception as e:
        print(f"  Erreur critique dans get_arp_table pour {net_connect.host}: {e}"); traceback.print_exc(); return []


def check_security_features(net_connect, running_config):
    security_audit = {}
    virtual_prefixes = ("vl", "lo", "tu", "po", "bv", "nu", "oo", "gr")

    # --- I. AAA & Authentification ---
    if "aaa new-model" in running_config:
        security_audit["aaa_new_model"] = {"status": True, "level": "good",
                                           "details": "AAA new-model est activé (prérequis pour TACACS+/RADIUS)."}
    else:
        security_audit["aaa_new_model"] = {"status": False, "level": "bad",
                                           "details": "AAA new-model n'est pas activé. Crucial pour la gestion centralisée et sécurisée des accès."}

    if "enable secret" in running_config:
        security_audit["enable_secret_configured"] = {"status": True, "level": "good",
                                                      "details": "Un 'enable secret' est configuré (hachage fort)."}
    elif "enable password" in running_config:
        security_audit["enable_secret_configured"] = {"status": "enable password only", "level": "bad",
                                                      "details": "'enable password' est utilisé sans 'enable secret'. Vulnérable même si 'service password-encryption' est actif."}
    else:
        security_audit["enable_secret_configured"] = {"status": False, "level": "bad",
                                                      "details": "Aucun 'enable secret' ou 'enable password' n'est configuré. Accès privilégié non protégé."}

    if "service password-encryption" in running_config:
        security_audit["password_encryption_service"] = {"status": True, "level": "good",
                                                         "details": "Service 'password-encryption' activé (obfusque mots de passe type 7, mais ne les sécurise pas fortement)."}
    else:
        security_audit["password_encryption_service"] = {"status": False, "level": "bad",
                                                         "details": "Service 'password-encryption' NON activé. Mots de passe (sauf 'enable secret') stockés en clair."}

    # --- II. Sécurité des Lignes d'Accès ---
    line_con_config_match = re.search(r"line con 0(.*?)!", running_config, re.DOTALL)
    con_config_text = line_con_config_match.group(1) if line_con_config_match else ""

    if "password" in con_config_text or "login local" in con_config_text or "login authentication" in con_config_text:
        security_audit["console_password"] = {"status": True, "level": "good",
                                              "details": "Ligne console protégée par méthode de login."}
    else:
        security_audit["console_password"] = {"status": False, "level": "bad",
                                              "details": "Ligne console non protégée par mot de passe."}

    if "exec-timeout" in con_config_text:
        timeout_match = re.search(r"exec-timeout\s+(\d+)\s*(?:(\d+))?", con_config_text)
        if timeout_match:
            minutes, secondes = int(timeout_match.group(1)), int(timeout_match.group(2) or 0)
            if minutes > 0 or (minutes == 0 and secondes > 0):
                security_audit["console_exec_timeout"] = {"status": f"{minutes}m {secondes}s", "level": "good",
                                                          "details": "Timeout d'exécution console configuré."}
            else:
                security_audit["console_exec_timeout"] = {"status": "Disabled (0 0)", "level": "bad",
                                                          "details": "Timeout d'exécution console désactivé (0 0)."}
        else:
            security_audit["console_exec_timeout"] = {"status": "Configured (check value)", "level": "warning",
                                                      "details": "exec-timeout console configuré, valeur à vérifier."}
    else:
        security_audit["console_exec_timeout"] = {"status": False, "level": "warning",
                                                  "details": "Aucun timeout d'exécution console. Recommandé : 5-15 minutes."}

    if "logging synchronous" in con_config_text:
        security_audit["console_logging_synchronous"] = {"status": True, "level": "good",
                                                         "details": "'logging synchronous' activé sur console."}
    else:
        security_audit["console_logging_synchronous"] = {"status": False, "level": "warning",
                                                         "details": "'logging synchronous' non activé sur console."}

    vty_config_text = ""
    vty_sections = re.findall(r"line vty\s+\d+\s*\d*(.*?)!", running_config, re.DOTALL)
    if vty_sections:
        vty_config_text = "\n".join(vty_sections)
    else:
        vty_config_text = running_config

    try:
        ssh_output_raw = net_connect.send_command("show ip ssh", use_textfsm=True)
        ssh_data = (ssh_output_raw[0] if isinstance(ssh_output_raw, list) and ssh_output_raw else {})
        ssh_version = ssh_data.get('protocol_version', 'N/A')
        if ssh_version == '2.0':
            security_audit["ssh_v2_only"] = {"status": True, "level": "good",
                                             "details": f"SSH version {ssh_version} activé et semble être la seule version."}
        elif ssh_version != 'N/A':
            security_audit["ssh_v2_only"] = {"status": False, "level": "bad",
                                             "details": f"SSH version {ssh_version} détectée. SSHv1 est activé et est vulnérable."}
        else:
            if "ip ssh version 2" in running_config and "no ip ssh version 1" in running_config:
                security_audit["ssh_v2_only"] = {"status": True, "level": "good",
                                                 "details": "SSH v2 explicitement configuré, v1 désactivé (config)."}
            elif "ip ssh version 2" in running_config:
                security_audit["ssh_v2_only"] = {"status": "v2 (v1 status unknown)", "level": "warning",
                                                 "details": "SSH v2 configuré, mais SSHv1 pourrait être encore actif. Ajouter 'no ip ssh version 1'."}
            elif "crypto key generate rsa" in running_config:
                security_audit["ssh_v2_only"] = {"status": "Unknown version", "level": "warning",
                                                 "details": "SSH semble activé (clés RSA), version non confirmée v2 uniquement."}
            else:
                security_audit["ssh_v2_only"] = {"status": False, "level": "bad",
                                                 "details": "SSH ne semble pas être activé ou configuré correctement."}
    except Exception as e:
        print(f"  Erreur SSH check pour {net_connect.host}: {e}"); security_audit["ssh_v2_only"] = {"status": "Error",
                                                                                                    "level": "warning",
                                                                                                    "details": "Vérif SSH impossible."}

    if "transport input telnet" in vty_config_text.lower():
        security_audit["vty_transport_telnet"] = {"status": "Telnet Enabled", "level": "bad",
                                                  "details": "Telnet autorisé sur lignes VTY. Protocole non chiffré, à désactiver."}
    elif "transport input ssh" in vty_config_text:
        security_audit["vty_transport_telnet"] = {"status": "SSH Only (Telnet not found)", "level": "good",
                                                  "details": "Lignes VTY semblent configurées pour SSH uniquement."}
    else:
        security_audit["vty_transport_telnet"] = {"status": "Transport Unclear", "level": "warning",
                                                  "details": "Configuration transport VTY non standard. Assurer SSH uniquement."}

    if "access-class" in vty_config_text and re.search(r"access-class\s+\S+\s+in", vty_config_text):
        security_audit["vty_acl"] = {"status": True, "level": "good",
                                     "details": "ACL (access-class) appliquée en entrée aux lignes VTY."}
    else:
        security_audit["vty_acl"] = {"status": False, "level": "bad",
                                     "details": "Aucune ACL (access-class) appliquée en entrée aux lignes VTY. Risque d'accès non filtré."}

    if "exec-timeout" in vty_config_text:
        vty_timeout_match = re.search(r"exec-timeout\s+(\d+)\s*(?:(\d+))?", vty_config_text)
        if vty_timeout_match:
            minutes, secondes = int(vty_timeout_match.group(1)), int(vty_timeout_match.group(2) or 0)
            if minutes > 0 or (minutes == 0 and secondes > 0):
                security_audit["vty_exec_timeout"] = {"status": f"{minutes}m {secondes}s", "level": "good",
                                                      "details": "Timeout d'exécution VTY configuré."}
            else:
                security_audit["vty_exec_timeout"] = {"status": "Disabled (0 0)", "level": "bad",
                                                      "details": "Timeout d'exécution VTY désactivé (0 0)."}
        else:
            security_audit["vty_exec_timeout"] = {"status": "Configured (check value)", "level": "warning",
                                                  "details": "exec-timeout VTY configuré, valeur à vérifier."}
    else:
        security_audit["vty_exec_timeout"] = {"status": False, "level": "warning",
                                              "details": "Aucun timeout d'exécution VTY. Recommandé : 5-15 minutes."}

    if "logging synchronous" in vty_config_text:
        security_audit["vty_logging_synchronous"] = {"status": True, "level": "good",
                                                     "details": "'logging synchronous' activé sur VTY."}
    else:
        security_audit["vty_logging_synchronous"] = {"status": False, "level": "warning",
                                                     "details": "'logging synchronous' non activé sur VTY."}

    # --- III. Renforcement Général ---
    if "no ip source-route" in running_config:
        security_audit["ip_source_route"] = {"status": "Disabled", "level": "good",
                                             "details": "Routage par la source désactivé."}
    else:
        security_audit["ip_source_route"] = {"status": "Enabled (default)", "level": "bad",
                                             "details": "Routage par la source actif. Configurer 'no ip source-route'."}

    if "no service finger" in running_config or "service finger" not in running_config:
        security_audit["finger_service"] = {"status": "Disabled", "level": "good",
                                            "details": "Service finger désactivé."}
    else:
        security_audit["finger_service"] = {"status": "Enabled", "level": "bad",
                                            "details": "Service finger activé. Configurer 'no service finger'."}

    small_servers_explicitly_disabled = "no service tcp-small-servers" in running_config and "no service udp-small-servers" in running_config
    small_servers_explicitly_enabled = "service tcp-small-servers" in running_config or "service udp-small-servers" in running_config
    if small_servers_explicitly_disabled or not small_servers_explicitly_enabled:
        security_audit["small_servers"] = {"status": "Disabled", "level": "good",
                                           "details": "Services tcp/udp-small-servers désactivés."}
    else:
        security_audit["small_servers"] = {"status": "Enabled", "level": "bad",
                                           "details": "tcp/udp-small-servers activés. À désactiver."}

    if "service timestamps log datetime msec" in running_config and "service timestamps debug datetime msec" in running_config:
        security_audit["service_timestamps"] = {"status": True, "level": "good",
                                                "details": "Horodatage précis (msec) logs/debug activé."}
    elif "service timestamps log" in running_config:
        security_audit["service_timestamps"] = {"status": "Partial", "level": "warning",
                                                "details": "Horodatage logs partiel. Recom: '... debug datetime msec' et '... log datetime msec'."}
    else:
        security_audit["service_timestamps"] = {"status": False, "level": "bad",
                                                "details": "Horodatage logs/debug non activé. Essentiel pour analyse."}

    if "banner motd" in running_config:
        security_audit["banner_motd"] = {"status": True, "level": "good", "details": "Bannière MOTD configurée."}
    else:
        security_audit["banner_motd"] = {"status": False, "level": "warning", "details": "Aucune bannière MOTD."}

    if "no ip http server" in running_config:
        security_audit["http_server"] = {"status": "Disabled", "level": "good",
                                         "details": "Serveur HTTP (non sécurisé) désactivé."}
    else:
        security_audit["http_server"] = {"status": "Enabled", "level": "bad",
                                         "details": "Serveur HTTP (non sécurisé) activé. Utiliser HTTPS ou désactiver."}

    if "ip http secure-server" in running_config:
        security_audit["https_server"] = {"status": "Enabled", "level": "good",
                                          "details": "Serveur HTTPS (sécurisé) activé."}
    elif "no ip http server" in running_config:
        security_audit["https_server"] = {"status": "Disabled (HTTP also disabled)", "level": "good",
                                          "details": "Serveur HTTPS désactivé (HTTP également désactivé)."}
    else:
        security_audit["https_server"] = {"status": "Disabled (HTTP is Enabled)", "level": "bad",
                                          "details": "Serveur HTTPS désactivé alors que HTTP est actif. Basculer vers HTTPS."}

    cdp_disabled, lldp_disabled = "no cdp run" in running_config, "no lldp run" in running_config
    security_audit["cdp_status"] = (
        {"status": "Globally Disabled", "level": "good", "details": "CDP désactivé globalement."} if cdp_disabled else
        {"status": "Globally Enabled", "level": "warning",
         "details": "CDP activé globalement. Filtrer sur interfaces non-confiance."})
    if lldp_disabled:
        security_audit["lldp_status"] = {"status": "Globally Disabled", "level": "good",
                                         "details": "LLDP désactivé globalement."}
    elif "lldp run" in running_config:
        security_audit["lldp_status"] = {"status": "Globally Enabled", "level": "warning",
                                         "details": "LLDP activé globalement. Filtrer sur interfaces non-confiance."}
    else:
        security_audit["lldp_status"] = {"status": "Potentially Enabled (default)", "level": "warning",
                                         "details": "LLDP potentiellement actif par défaut."}

    try:
        int_status_raw = net_connect.send_command("show interfaces status", use_textfsm=True)
        int_status = int_status_raw if isinstance(int_status_raw, list) else []
        unused_ports_details, admin_down_ports_details = [], []
        for i_data in int_status:
            port_name, port_status = i_data.get('port'), i_data.get('status', '').lower()
            if not port_name or any(port_name.lower().startswith(p) for p in virtual_prefixes): continue
            if port_status in ['notconnect', 'disabled']: unused_ports_details.append(port_name)
            if port_status == 'disabled': admin_down_ports_details.append(port_name)
        active_unused = [p for p in unused_ports_details if p not in admin_down_ports_details]
        if not active_unused and not unused_ports_details:
            security_audit["unused_physical_ports"] = {"status": "All connected or N/A", "level": "good",
                                                       "details": "Tous les ports physiques sont connectés ou statut non détaillé."}
        elif not active_unused:
            security_audit["unused_physical_ports"] = {"status": f"{len(admin_down_ports_details)} désactivés",
                                                       "level": "good",
                                                       "details": f"Ports physiques inutilisés et désactivés: {', '.join(admin_down_ports_details)}."}
        else:
            security_audit["unused_physical_ports"] = {"status": f"{len(active_unused)} actifs non connectés",
                                                       "level": "bad",
                                                       "details": f"Ports physiques non connectés mais actifs: {', '.join(active_unused)}. Risque. Mettre en 'shutdown'."}
    except Exception as e:
        print(f"  Erreur unused_ports check pour {net_connect.host}: {e}"); security_audit["unused_physical_ports"] = {
            "status": "Error", "level": "warning", "details": f"Vérif ports inutilisés impossible: {e}"}

    # --- IV. Logging & Monitoring ---
    if "logging buffered" in running_config:
        buflvl_match = re.search(r"logging buffered\s+(?:\d+|\S+)", running_config)
        if buflvl_match:
            security_audit["logging_buffered"] = {"status": f"Activé ({buflvl_match.group(0).split()[-1]})",
                                                  "level": "good", "details": "Logging bufferisé activé."}
        else:
            security_audit["logging_buffered"] = {"status": "Activé (defaults)", "level": "good",
                                                  "details": "Logging bufferisé activé (défauts)."}
    else:
        security_audit["logging_buffered"] = {"status": False, "level": "warning",
                                              "details": "Logging bufferisé non activé."}

    if "logging host" in running_config or "logging server" in running_config:
        security_audit["remote_logging_configured"] = {"status": True, "level": "good",
                                                       "details": "Logging distant (syslog) configuré."}
        if "logging source-interface" in running_config:
            src_int_match = re.search(r"logging source-interface\s+(\S+)", running_config)
            security_audit["logging_source_interface"] = {
                "status": src_int_match.group(1) if src_int_match else "Configurée", "level": "good",
                "details": "Interface source syslog spécifiée."}
        else:
            security_audit["logging_source_interface"] = {"status": False, "level": "warning",
                                                          "details": "Aucune interface source syslog."}
    else:
        security_audit["remote_logging_configured"] = {"status": False, "level": "bad",
                                                       "details": "Logging distant (syslog) NON configuré."}
        security_audit["logging_source_interface"] = {"status": "N/A", "level": "bad",
                                                      "details": "Syslog non configuré."}

    ntp_servers = [line for line in running_config.splitlines() if
                   "ntp server " in line.strip() and not line.strip().startswith("ntp server vrf")]
    num_ntp_servers = len(ntp_servers)
    ntp_sync_status, ntp_sync_details, ntp_sync_level = "Error", "Statut synchro NTP inconnu.", "warning"
    try:
        ntp_status_raw = net_connect.send_command("show ntp status", use_textfsm=True)
        ntp_data = (ntp_status_raw[0] if isinstance(ntp_status_raw, list) and ntp_status_raw else {})
        clock_state = ntp_data.get('clock_state', '').lower()
        if "synchronised" in clock_state or "synchronized" in clock_state:
            ntp_sync_status, ntp_sync_details, ntp_sync_level = True, f"NTP synchronisé. Stratum: {ntp_data.get('stratum', 'N/A')}, Serveur ref: {ntp_data.get('reference_server', 'N/A')}.", "good"
        else:
            ntp_sync_status, ntp_sync_details = False, f"NTP non synchronisé (état: {clock_state}). Logs incorrectement horodatés."
            ntp_sync_level = "bad"
    except Exception:
        ntp_sync_details = (
            "NTP configuré, statut synchro inconnu ('show ntp status' échoué)." if num_ntp_servers > 0 else
            "NTP non configuré et statut inconnu.")
        ntp_sync_level = "warning" if num_ntp_servers > 0 else "bad"
    security_audit["ntp_synchronization"] = {"status": ntp_sync_status, "level": ntp_sync_level,
                                             "details": ntp_sync_details}

    if num_ntp_servers >= 2:
        security_audit["ntp_redundancy"] = {"status": f"{num_ntp_servers} serveurs", "level": "good",
                                            "details": "Redondance NTP OK."}
    elif num_ntp_servers == 1:
        security_audit["ntp_redundancy"] = {"status": "1 serveur", "level": "warning",
                                            "details": "Un seul serveur NTP. Recom: >=2."}
    else:
        security_audit["ntp_redundancy"] = {"status": "0 serveur", "level": "bad",
                                            "details": "Aucun serveur NTP configuré. Temps non fiable."}

    # --- V. Sécurité Couche 2 (Indications) ---
    if "switchport port-security" in running_config:
        security_audit["port_security_feature_used"] = {"status": True, "level": "good",
                                                        "details": "Fonctionnalité Port Security utilisée."}
    else:
        security_audit["port_security_feature_used"] = {"status": False, "level": "bad",
                                                        "details": "Port Security non utilisé. Essentiel sur ports d'accès."}

    if "ip dhcp snooping" in running_config:
        security_audit["dhcp_snooping_global"] = {"status": True, "level": "good",
                                                  "details": "DHCP Snooping activé globalement."}
    else:
        security_audit["dhcp_snooping_global"] = {"status": False, "level": "bad",
                                                  "details": "DHCP Snooping non activé globalement. Requis pour prévenir serveurs DHCP pirates."}

    if "spanning-tree portfast bpduguard default" in running_config:
        security_audit["bpduguard_default"] = {"status": True, "level": "good",
                                               "details": "BPDU Guard activé par défaut sur ports PortFast."}
    elif "spanning-tree bpduguard enable" in running_config:
        security_audit["bpduguard_default"] = {"status": "Per-interface (check)", "level": "good",
                                               "details": "BPDU Guard activé sur certaines interfaces."}
    else:
        security_audit["bpduguard_default"] = {"status": False, "level": "warning",
                                               "details": "BPDU Guard non activé par défaut. Risque de boucles STP."}

    if "storm-control" in running_config:
        security_audit["storm_control_used"] = {"status": True, "level": "good",
                                                "details": "Storm Control semble configuré."}
    else:
        security_audit["storm_control_used"] = {"status": False, "level": "warning",
                                                "details": "Storm Control non utilisé."}

    if "snmp-server community public RO" in running_config or "snmp-server community private RW" in running_config:
        security_audit["snmp_default_communities"] = {"status": True, "level": "bad",
                                                      "details": "SNMP utilise des communautés PAR DÉFAUT. Risque majeur."}
    elif "snmp-server community" in running_config:
        security_audit["snmp_default_communities"] = {"status": False, "level": "good",
                                                      "details": "SNMP configuré (pas de communautés par défaut)."}
    else:
        security_audit["snmp_default_communities"] = {"status": "Not Configured", "level": "good",
                                                      "details": "SNMP (v1/v2c community) non configuré."}

    if "snmp-server group" in running_config and "v3 auth" in running_config:
        security_audit["snmp_version"] = {"status": "v3 (probable)", "level": "good",
                                          "details": "SNMPv3 semble utilisé."}
    elif "snmp-server community" in running_config:
        security_audit["snmp_version"] = {"status": "v1/v2c", "level": "bad",
                                          "details": "SNMPv1/v2c utilisé (communautés en clair). Préférer SNMPv3."}
    else:
        security_audit["snmp_version"] = {"status": "N/A", "level": "good", "details": "SNMP non configuré."}
    return security_audit


def load_inventory(filepath="inventory.csv"):
    inventory = []
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
                if len(header) < 3:
                    print(
                        f"Erreur: En-tête inventaire '{filepath}' doit avoir au moins 3 colonnes (hostname, group, device_type).")
                    return None
            except StopIteration:
                print(f"Avertissement: Fichier inventaire '{filepath}' vide."); return []
            for row in reader:
                if len(row) >= 3 and row[0].strip():
                    inventory.append(
                        {"host": row[0].strip(), "group": row[1].strip(), "device_type": row[2].strip().lower()})
                elif row and any(field.strip() for field in row):
                    print(f"Avertissement: Ligne inventaire mal formatée: {row}")
        return inventory
    except FileNotFoundError:
        print(f"Erreur: Fichier inventaire '{filepath}' non trouvé."); return None
    except Exception as e:
        print(f"Erreur lecture '{filepath}': {e}"); traceback.print_exc(); return None


def load_passwords(filepath="passwords.csv"):
    passwords = {}
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            try:
                next(reader)
            except StopIteration:
                print(f"Avertissement: Fichier passwords '{filepath}' vide."); return {}
            for row in reader:
                if len(row) >= 3 and row[0].strip():
                    enable_pass = row[3].strip() if len(row) > 3 and row[3].strip() else None
                    passwords[row[0].strip()] = {"username": row[1].strip(), "password": row[2].strip(),
                                                 "enable_password": enable_pass}
                elif row and any(field.strip() for field in row):
                    print(f"Avertissement: Ligne passwords mal formatée: {row}")
        return passwords
    except FileNotFoundError:
        print(f"Erreur: Fichier passwords '{filepath}' non trouvé."); return None
    except Exception as e:
        print(f"Erreur lecture '{filepath}': {e}"); traceback.print_exc(); return None


def generate_excel_report(all_data, excel_filepath):  # Modifié pour prendre un chemin complet
    if not all_data: print("Aucune donnée pour rapport Excel."); return
    wb = openpyxl.Workbook();
    wb.remove(wb.active)
    ws_info = wb.create_sheet("Infos Générales")
    headers_info = ["Hostname", "IP Address", "Modèle", "Version IOS", "Uptime", "Numéro de Série"]
    ws_info.append(headers_info);
    apply_header_style(ws_info)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection':
            ws_info.append(
                [dev_data.get('attempted_host', 'N/A'), dev_data.get('attempted_host', 'N/A'), "ERREUR CONNEXION",
                 dev_data.get('error_message', 'N/A'), "N/A", "N/A"])
            for i in range(3, 5): ws_info.cell(row=ws_info.max_row, column=i).fill = RED_FILL
        else:
            info = dev_data.get("general_info", {})
            ws_info.append([info.get("hostname", dev_data.get("host", "N/A")), dev_data.get("host", "N/A"),
                            info.get("model", "N/A"), info.get("ios_version", "N/A"), info.get("uptime", "N/A"),
                            info.get("serial_number", "N/A")])
    auto_fit_columns(ws_info)
    ws_interfaces = wb.create_sheet("Interfaces")
    headers_interfaces = ["Hostname", "Nom Interface", "Type", "Description", "IP Address", "Statut Lien",
                          "Statut Protocole", "VLAN (Access)", "Duplex", "Vitesse"]
    ws_interfaces.append(headers_interfaces);
    apply_header_style(ws_interfaces)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for iface in dev_data.get("interfaces", []):
            ws_interfaces.append(
                [hostname, iface.get("name"), iface.get("type"), iface.get("description"), iface.get("ip_address"),
                 iface.get("status_link"), iface.get("status_protocol"), iface.get("vlan"), iface.get("duplex"),
                 iface.get("speed")])
            lr, pr = ws_interfaces.cell(row=ws_interfaces.max_row, column=6), ws_interfaces.cell(
                row=ws_interfaces.max_row, column=7)
            sl, sp = str(iface.get("status_link", "")).lower(), str(iface.get("status_protocol", "")).lower()
            if sl == "up" or "connected" in sl:
                lr.fill = GREEN_FILL
            elif ("admin" in sl and "down" in sl) or sl == "disabled":
                lr.fill = ORANGE_FILL
            elif sl == "down" or "notconnect" in sl or "err-disabled" in sl or "error-disabled" in sl:
                lr.fill = RED_FILL
            if sp == "up":
                pr.fill = GREEN_FILL
            elif sp == "down":
                pr.fill = RED_FILL
    auto_fit_columns(ws_interfaces)
    ws_vlans = wb.create_sheet("VLANs")
    headers_vlans = ["Hostname", "ID VLAN", "Nom VLAN", "Statut", "Ports Assignés"]
    ws_vlans.append(headers_vlans);
    apply_header_style(ws_vlans)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for vlan in dev_data.get("vlans", []):
            ws_vlans.append([hostname, vlan.get("id"), vlan.get("name"), vlan.get("status"), vlan.get("ports")])
            sc = ws_vlans.cell(row=ws_vlans.max_row, column=4)
            if str(vlan.get("status", "")).lower() == "active":
                sc.fill = GREEN_FILL
            else:
                sc.fill = ORANGE_FILL
    auto_fit_columns(ws_vlans)
    ws_arp = wb.create_sheet("Table ARP")
    headers_arp = ["Hostname", "Protocole", "Adresse IP", "Âge (min)", "Adresse MAC", "Type", "Interface"]
    ws_arp.append(headers_arp);
    apply_header_style(ws_arp)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for entry in dev_data.get("arp_table", []):
            ws_arp.append(
                [hostname, entry.get("protocol"), entry.get("address"), entry.get("age"), entry.get("mac_address"),
                 entry.get("type"), entry.get("interface")])
    auto_fit_columns(ws_arp)
    ws_security = wb.create_sheet("Audit Sécurité")
    headers_security = ["Hostname", "Point de Contrôle", "Statut/Valeur", "Niveau", "Détails/Recommandation"]
    ws_security.append(headers_security);
    apply_header_style(ws_security)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for check_name, check_data in dev_data.get("security_audit", {}).items():
            ws_security.append(
                [hostname, check_name.replace("_", " ").title(), str(check_data.get("status")), check_data.get("level"),
                 check_data.get("details")])
            lc = ws_security.cell(row=ws_security.max_row, column=4)
            set_cell_status_color(lc, check_data.get("level"))
    auto_fit_columns(ws_security)
    # excel_filepath est maintenant passé en argument
    try:
        wb.save(excel_filepath); print(f"\n[+] Rapport Excel généré : {excel_filepath}")
    except Exception as e:
        print(f"\n[-] Erreur sauvegarde Excel: {e}")


def perform_cisco_audit(cisco_devices_inventory, global_passwords_map, output_directory):
    all_devices_data = []
    for device_entry in cisco_devices_inventory:
        host, group = device_entry["host"], device_entry["group"]
        creds = global_passwords_map.get(group)
        print(f"\n[INFO Cisco] Traitement de {host} (groupe: {group})...")
        if not creds:
            print(f"  [ERREUR Cisco] Identifiants non trouvés pour groupe '{group}'. {host} ignoré.")
            all_devices_data.append({"attempted_host": host, "status": "error_connection",
                                     "error_message": f"Identifiants non trouvés pour groupe {group}"})
            continue
        dev_params = {'device_type': 'cisco_ios', 'host': host, 'username': creds['username'],
                      'password': creds['password'],
                      'secret': creds.get('enable_password'), 'global_delay_factor': 2, 'timeout': 45,
                      'session_timeout': 120}
        current_device_data = {"host": host}
        try:
            with ConnectHandler(**dev_params) as net_connect:
                actual_host, actual_prompt = net_connect.host, (
                    net_connect.base_prompt[:-1] if net_connect.base_prompt else net_connect.host)
                print(f"  [OK Cisco] Connecté à {actual_host} ({actual_prompt}).")
                if creds.get('enable_password'): net_connect.enable()

                current_device_data["general_info"] = get_device_info(net_connect)
                current_device_data["general_info"]["ip_address_queried"] = host

                running_config = net_connect.send_command("show running-config", read_timeout=240)
                if not running_config: running_config = ""

                current_device_data["interfaces"] = get_interfaces(net_connect)
                current_device_data["vlans"] = get_vlans(net_connect)
                current_device_data["arp_table"] = get_arp_table(net_connect)
                current_device_data["security_audit"] = check_security_features(net_connect, running_config)
                all_devices_data.append(current_device_data)
        except (NetmikoTimeoutException, SSHException) as e:
            print(f"  [ERREUR Cisco] Connexion à {host} (Timeout/SSH): {e}")
            all_devices_data.append({"attempted_host": host, "status": "error_connection", "error_message": str(e)})
        except NetmikoAuthenticationException as e:
            print(f"  [ERREUR Cisco] Authentification sur {host}: {e}")
            all_devices_data.append(
                {"attempted_host": host, "status": "error_connection", "error_message": f"Échec authentification: {e}"})
        except Exception as e:
            print(f"  [ERREUR Cisco] Inattendue avec {host}: {e}");
            traceback.print_exc()
            all_devices_data.append(
                {"attempted_host": host, "status": "error_connection", "error_message": f"Erreur inattendue: {e}"})

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = os.path.join(output_directory, f"audit_cisco_data_{timestamp}.json")
    try:
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(all_devices_data, f, indent=4, ensure_ascii=False)
        print(f"\n[+] Données JSON Cisco sauvegardées : {json_filename}")
    except Exception as e:
        print(f"\n[-] Erreur sauvegarde JSON Cisco: {e}")

    excel_report_path = os.path.join(output_directory, f"audit_cisco_report_{timestamp}.xlsx")
    generate_excel_report(all_devices_data, excel_report_path)
    print(f"\n[+] Audit Cisco terminé pour {len(cisco_devices_inventory)} équipement(s).")


def main():
    inventory_file, password_file, output_directory = "inventory.csv", "passwords.csv", "audit_reports"
    if not os.path.exists(output_directory):
        try:
            os.makedirs(output_directory)
        except OSError as e:
            print(f"Erreur création répertoire '{output_directory}': {e}"); return

    full_inventory = load_inventory(inventory_file)
    passwords_map = load_passwords(password_file)

    if full_inventory is None or passwords_map is None:
        print("Arrêt (Cisco): Erreurs critiques chargement fichiers entrée.");
        return

    cisco_devices = [device for device in full_inventory if device.get("device_type") in ["cisco_ios", "cisco_iosxe"]]

    if not cisco_devices:
        print("Aucun équipement Cisco (cisco_ios/cisco_iosxe) trouvé dans l'inventaire pour audit autonome.")
        return

    perform_cisco_audit(cisco_devices, passwords_map, output_directory)


if __name__ == "__main__":
    main()