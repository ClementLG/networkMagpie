# aruba_audit.py
# !/usr/bin/env python3

# NetworkMagpie
# Copyright (C) 2025 CLEMENT LE GRUIEC
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import json
import csv
import os
import datetime
import re
from netmiko import ConnectHandler
from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException, SSHException
from textfsm.parser import TextFSMError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import traceback

# --- Constantes Excel et fonctions d'aide ---
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BOLD_FONT = Font(bold=True)


def apply_header_style(ws, row_num=1):
    for cell in ws[row_num]:
        cell.font = BOLD_FONT;
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0;
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)


def set_cell_status_color(cell, level):
    if level == "good":
        cell.fill = GREEN_FILL
    elif level == "warning":
        cell.fill = ORANGE_FILL
    elif level == "error" or level == "bad":
        cell.fill = RED_FILL


# --- Fonctions de collecte Aruba OS-CX ---
def get_aruba_device_info(net_connect):
    info = {'hostname': 'N/A', 'ios_version': 'N/A', 'model': 'N/A', 'serial_number': 'N/A', 'uptime': 'N/A'}
    raw_sys_data = ""  # Pour stocker la sortie brute de 'show system' si nécessaire
    try:
        prompt_hostname = net_connect.base_prompt
        if prompt_hostname:
            cleaned_prompt = re.sub(r"[#>()\s]+$", "", prompt_hostname)
            cleaned_prompt = cleaned_prompt.splitlines()[-1]
            if cleaned_prompt: info['hostname'] = cleaned_prompt

        # 1. Tenter 'show system' avec TextFSM
        system_out_textfsm = None
        try:
            system_out_textfsm = net_connect.send_command("show system", use_textfsm=True, expect_string=r"#")
            if system_out_textfsm and isinstance(system_out_textfsm, list) and system_out_textfsm[0]:
                sys_data_textfsm = system_out_textfsm[0]
                info['hostname'] = sys_data_textfsm.get('hostname', info['hostname'])
                info['model'] = sys_data_textfsm.get('product_name', sys_data_textfsm.get('model', 'N/A'))
                info['serial_number'] = sys_data_textfsm.get('chassis_serial_nbr',
                                                             sys_data_textfsm.get('serial_number', 'N/A'))
                if sys_data_textfsm.get('up_time') and sys_data_textfsm.get('up_time', 'N/A') != 'N/A':
                    info['uptime'] = sys_data_textfsm['up_time'].strip()
            # else: # TextFSM a retourné quelque chose de non conforme, on passera au parsing brut
            # print(f"  [AVERTISSEMENT] TextFSM 'show system' sur {net_connect.host} n'a pas retourné les données attendues.")
        except (TextFSMError, ValueError, IndexError) as e:
            print(
                f"  [AVERTISSEMENT] TextFSM pour 'show system' sur {net_connect.host} a échoué: {e}. Passage en mode brut.")

        # 2. Si des infos manquent après TextFSM pour 'show system', utiliser le parsing brut
        if info['model'] == 'N/A' or info['serial_number'] == 'N/A' or info['uptime'] == 'N/A':
            if not raw_sys_data:  # Récupérer la sortie brute si ce n'est pas déjà fait
                raw_sys_data = net_connect.send_command("show system", use_textfsm=False, expect_string=r"#")
            # print(f"DEBUG: Host {net_connect.host} - raw 'show system' (pour fallback info): \n{raw_sys_data}\n--------------------")

            if info['model'] == 'N/A':
                match_model_sys = re.search(r"Product Name\s*:\s*([^\r\n]+)", raw_sys_data, re.IGNORECASE)
                if match_model_sys: info['model'] = match_model_sys.group(1).strip()

            if info['serial_number'] == 'N/A':
                match_serial_sys = re.search(r"Chassis Serial Nbr\s*:\s*(\S+)", raw_sys_data, re.IGNORECASE)
                if match_serial_sys: info['serial_number'] = match_serial_sys.group(1).strip()

            if info['uptime'] == 'N/A':  # Uptime spécifique de 'show system'
                match_uptime_sys = re.search(r"Up Time\s*:\s*(.+)", raw_sys_data, re.IGNORECASE)
                if match_uptime_sys: info['uptime'] = match_uptime_sys.group(1).strip()

        # 3. Obtenir la version OS depuis 'show version' (plus spécifique pour la version)
        version_out_textfsm = None
        try:
            version_out_textfsm = net_connect.send_command("show version", use_textfsm=True, expect_string=r"#")
            if version_out_textfsm and isinstance(version_out_textfsm, list) and version_out_textfsm[0]:
                ver_data = version_out_textfsm[0]
                if info['hostname'] == 'N/A' and ver_data.get('hostname'): info['hostname'] = ver_data.get('hostname')
                info['ios_version'] = ver_data.get('version',
                                                   ver_data.get('software_version',
                                                                ver_data.get('os_version',
                                                                             ver_data.get('arubaos_cx_version',
                                                                                          'N/A'))))
            else:
                raise ValueError("TextFSM pour 'show version' n'a pas retourné de données valides.")
        except (TextFSMError, ValueError, IndexError) as e:
            # print(f"  [AVERTISSEMENT] TextFSM/Parsing 'show version' sur {net_connect.host} a échoué: {e}. Mode brut.") # Logged par console
            raw_ver_data = net_connect.send_command("show version", use_textfsm=False, expect_string=r"#")
            if info['hostname'] == 'N/A':
                match_hostname_ver = re.search(r"Hostname\s*:\s*(\S+)", raw_ver_data, re.IGNORECASE)
                if match_hostname_ver: info['hostname'] = match_hostname_ver.group(1)
            match_version_ver = re.search(r"(?:ArubaOS-CX Version|Software version|Version)\s*:\s*(\S+)", raw_ver_data,
                                          re.IGNORECASE)
            if match_version_ver: info['ios_version'] = match_version_ver.group(1)

        # 4. Si l'uptime n'a toujours pas été trouvé, essayer 'show uptime' comme dernier recours
        if info.get('uptime', 'N/A') == 'N/A':
            uptime_out_raw = net_connect.send_command("show uptime", expect_string=r"#", use_textfsm=False)
            match_uptime_cmd = re.search(r"(?:System Uptime|Uptime is)\s*:\s*(.+)", uptime_out_raw, re.IGNORECASE)
            if match_uptime_cmd: info['uptime'] = match_uptime_cmd.group(1).strip()

        return info
    except Exception as e:
        print(f"  Erreur critique get_aruba_device_info: {e}")
        traceback.print_exc()
        return info  # Retourner ce qui a été collecté jusqu'à présent


def get_aruba_interfaces(net_connect):
    interfaces = []
    try:
        ip_brief_textfsm_out = net_connect.send_command("show ip interface brief", use_textfsm=True, expect_string=r"#")
        ip_map = {}
        if isinstance(ip_brief_textfsm_out, list):
            for item in ip_brief_textfsm_out:
                ifname = item.get('interface', item.get('intf'))
                if ifname:
                    ip_addr = item.get('ip_address', item.get('ipaddr'))
                    protocol_l3 = item.get('protocol', 'N/A').lower()
                    status_l3 = item.get('status', 'N/A').lower()
                    entry = {"protocol_l3": protocol_l3, "status_l3": status_l3}
                    if ip_addr and ip_addr != 'N/A' and ip_addr != 'unassigned':
                        entry["ip"] = ip_addr
                    else:
                        entry["ip"] = "unassigned"
                    ip_map[ifname] = entry

        int_status_list = []
        try:
            int_status_out = net_connect.send_command("show interface status", use_textfsm=True, expect_string=r"#")
            if isinstance(int_status_out, list): int_status_list = int_status_out
        except TextFSMError:
            pass
        except Exception as e:
            print(f"  [ERREUR] Commande 'show interface status' {net_connect.host}: {e}.")

        status_l2_map = {item.get('port', item.get('interface')): item for item in int_status_list if
                         item.get('port', item.get('interface'))}

        running_config_interfaces = net_connect.send_command("show running-config interfaces", read_timeout=120,
                                                             expect_string=r"#")
        desc_map = {}
        current_if_desc = None
        for line in running_config_interfaces.splitlines():
            line_strip = line.strip()
            if line_strip.startswith("interface "):
                current_if_desc = line_strip.split()[-1]
            elif "description " in line_strip and current_if_desc:
                desc_map[current_if_desc] = line_strip.split("description ", 1)[1]
            elif not line_strip or line_strip.startswith("!"):
                current_if_desc = None

        show_int_brief_raw = net_connect.send_command("show interface brief", use_textfsm=False, expect_string=r"#")

        int_brief_re = re.compile(
            r"^(?P<interface>\S+)\s+"
            r"(?P<native_vlan>\S+)\s+"
            r"(?P<mode>\S+)\s+"
            r"(?P<type_col>\S+)\s+"
            r"(?P<enabled>yes|no)\s+"
            r"(?P<link_status_l2>\S+)\s+"
            r"(?P<reason>.*?)\s{2,}"
            r"(?P<speed>\S+)\s+"
            r"(?P<description>.*)$"
        )
        int_brief_re_short = re.compile(
            r"^(?P<interface>\S+)\s+"
            r"(?P<native_vlan>\S+)\s+"
            r"(?P<mode>\S+)\s+"
            r"(?P<type_col>\S+)\s+"
            r"(?P<enabled>yes|no)\s+"
            r"(?P<link_status_l2>\S+)\s+"
            r"(?P<speed>\S+)\s+"
            r"(?P<description>.*)$"
        )

        interfaces_from_regex_count = 0
        header_skipped_brief = False
        processed_interfaces_in_brief = set()

        for line in show_int_brief_raw.splitlines():
            line_s = line.strip()
            if not line_s: continue
            if line_s.lower().startswith("port ") or \
                    line_s.lower().startswith("native") or \
                    line_s.lower().startswith("-----"):
                header_skipped_brief = True;
                continue
            if not header_skipped_brief: continue

            match = int_brief_re.match(line_s)
            if not match: match = int_brief_re_short.match(line_s)

            if match:
                interfaces_from_regex_count += 1
                data = match.groupdict()
                name = data['interface']
                processed_interfaces_in_brief.add(name)

                ip_info_dict = ip_map.get(name, {})
                ip = ip_info_dict.get("ip", "N/A")
                final_proto_status = ip_info_dict.get("protocol_l3", "N/A")

                status_l2_detail = status_l2_map.get(name, {})

                admin_enabled = data.get('enabled', 'no').lower()
                link_s_l2 = data.get('link_status_l2', 'N/A').lower()
                reason = data.get('reason', '').strip().lower()

                final_link_status = link_s_l2
                if admin_enabled == 'no':
                    final_link_status = "administratively down"
                elif reason == "administratively down":
                    final_link_status = "administratively down"
                elif reason == "no xcvr installed":
                    final_link_status = "down (no transceiver)"

                link_s_from_status_cmd = status_l2_detail.get('status', '').lower()
                if link_s_from_status_cmd: final_link_status = link_s_from_status_cmd

                status_l3_from_ip = ip_info_dict.get("status_l3", "N/A")
                if final_proto_status == 'n/a' or final_proto_status == '':
                    if status_l3_from_ip != 'n/a' and status_l3_from_ip != '':
                        final_proto_status = status_l3_from_ip
                    elif final_link_status == "up":
                        final_proto_status = "up"
                    elif final_link_status == "administratively down":
                        final_proto_status = "down"
                    elif final_link_status.startswith("down"):
                        final_proto_status = "down"

                intf_type_parsed = "Management" if name.lower() == "mgmt" else \
                    "Virtual" if name.lower().startswith(("vlan", "loopback", "lag")) else \
                        "Physical"

                vlan_info = status_l2_detail.get('vlan', 'N/A')
                if data.get('native_vlan') and data.get('native_vlan') != '--':
                    vlan_info = data.get('native_vlan')
                if name.lower().startswith("vlan") and name[4:].isdigit():
                    vlan_info = name[4:]

                interfaces.append({
                    "name": name, "ip_address": ip,
                    "status_link": final_link_status, "status_protocol": final_proto_status,
                    "description": desc_map.get(name, data.get('description', "N/A").strip()),
                    "type": intf_type_parsed,
                    "vlan": vlan_info,
                    "duplex": status_l2_detail.get('duplex', 'N/A'),
                    "speed": status_l2_detail.get('speed', data.get('speed', 'N/A')),
                })

        # Traitement spécifique pour l'interface mgmt si elle n'a pas été trouvée
        if "mgmt" not in processed_interfaces_in_brief:
            try:
                mgmt_raw = net_connect.send_command("show interface mgmt", use_textfsm=False, expect_string=r"#")
                # print(f"DEBUG: Host {net_connect.host} - raw 'show interface mgmt':\n{mgmt_raw}\n--------------------")

                mgmt_ip, mgmt_link, mgmt_admin, mgmt_proto = "N/A", "N/A", "N/A", "N/A"

                match_ip = re.search(r"IPv4 address/subnet-mask\s*:\s*(\S+)", mgmt_raw, re.IGNORECASE)
                if match_ip: mgmt_ip = match_ip.group(1)

                match_admin = re.search(r"Admin State\s*:\s*(\S+)", mgmt_raw, re.IGNORECASE)
                if match_admin: mgmt_admin = match_admin.group(1).lower()

                match_link = re.search(r"Link State\s*:\s*(\S+)", mgmt_raw, re.IGNORECASE)
                if match_link: mgmt_link = match_link.group(1).lower()

                if mgmt_admin == "down":
                    final_mgmt_link_status = "administratively down"
                else:
                    final_mgmt_link_status = mgmt_link

                if final_mgmt_link_status == "up":
                    mgmt_proto = "up"
                elif final_mgmt_link_status == "administratively down":
                    mgmt_proto = "down"
                elif final_mgmt_link_status == "down":
                    mgmt_proto = "down"

                # Si l'interface mgmt a une IP via 'show ip interface brief', cela sera prioritaire
                mgmt_ip_from_map = ip_map.get("mgmt", {}).get("ip", mgmt_ip)
                mgmt_proto_from_map = ip_map.get("mgmt", {}).get("protocol_l3", mgmt_proto)
                if mgmt_proto_from_map != 'N/A': mgmt_proto = mgmt_proto_from_map

                if mgmt_link != "N/A":  # Si on a pu lire quelque chose de 'show interface mgmt'
                    interfaces.append({
                        "name": "mgmt", "ip_address": mgmt_ip_from_map,
                        "status_link": final_mgmt_link_status, "status_protocol": mgmt_proto,
                        "description": desc_map.get("mgmt", "Management Interface"), "type": "Management",
                        "vlan": "N/A", "duplex": "N/A", "speed": "N/A",  # Non dispo facilement via cette commande
                    })
                    interfaces_from_regex_count += 1  # Compter l'interface mgmt
            except Exception as e_mgmt:
                print(
                    f"  [AVERTISSEMENT] Erreur lors de la récupération de 'show interface mgmt' sur {net_connect.host}: {e_mgmt}")

        print(
            f"  [INFO] Parsing des interfaces a traité/trouvé {interfaces_from_regex_count} entrées pour {net_connect.host}.")
        return interfaces
    except Exception as e:
        print(f"  Erreur critique dans get_aruba_interfaces pour {net_connect.host}: {e}");
        traceback.print_exc()
        return []


def get_vlans(net_connect):
    vlans_list = []
    try:
        vlan_out_textfsm = None
        try:
            vlan_out_textfsm = net_connect.send_command("show vlan", use_textfsm=True, expect_string=r"#")
        except TextFSMError:
            pass
        except Exception as e:
            print(f"  [ERREUR] Commande 'show vlan' (TextFSM) a échoué sur {net_connect.host}: {e}.")

        vlan_data_textfsm = vlan_out_textfsm if isinstance(vlan_out_textfsm, list) else []

        if vlan_data_textfsm:
            for v_entry in vlan_data_textfsm:
                vlans_list.append({
                    "id": v_entry.get('vlan_id', v_entry.get('id', 'N/A')),
                    "name": v_entry.get('name', v_entry.get('vlan_name', 'N/A')),
                    "status": v_entry.get('status', 'N/A'),
                    "ports": ", ".join(v_entry.get('ports', [])) if v_entry.get('ports') and isinstance(
                        v_entry.get('ports'), list) else v_entry.get('ports', 'N/A')
                })
        else:
            raw_vlan_out = net_connect.send_command("show vlan", use_textfsm=False, expect_string=r"#")
            header_skipped = False
            for line in raw_vlan_out.splitlines():
                line_s = line.strip()
                if not line_s: continue
                if line_s.startswith("----") or line_s.lower().startswith("vlan name") or line_s.lower().startswith(
                        "vlan id name"):
                    header_skipped = True;
                    continue
                if not header_skipped: continue
                match_vlan = re.match(r"^\s*(\d+)\s+([\w\-\/\.]+)\s+(\S+)(?:\s+\S*){2}\s*(.*)", line_s)
                if match_vlan:
                    vid, vname, vstatus, vports_str = match_vlan.groups()
                    vname_clean = vname if not vname.startswith("<NO") else "N/A"
                    vports_clean = vports_str.strip() if vports_str.strip() and not vports_str.startswith(
                        "<NO") else "N/A"
                    vlans_list.append({"id": vid, "name": vname_clean, "status": vstatus, "ports": vports_clean})
        return vlans_list
    except Exception as e:
        print(f"  Erreur critique dans get_vlans pour {net_connect.host}: {e}");
        traceback.print_exc()
        return []


def get_arp_table(net_connect):
    arp_table = []
    try:
        arp_out_textfsm = None
        try:
            arp_out_textfsm = net_connect.send_command("show arp", use_textfsm=True, expect_string=r"#")
        except TextFSMError:
            pass
        except Exception as e:
            print(f"  [ERREUR] Commande 'show arp' (TextFSM) a échoué sur {net_connect.host}: {e}.")

        arp_data_textfsm = arp_out_textfsm if isinstance(arp_out_textfsm, list) else []

        if arp_data_textfsm:
            for entry in arp_data_textfsm:
                arp_table.append({
                    "protocol": entry.get('protocol', 'Internet'),
                    "address": entry.get('address', entry.get('ip_address', 'N/A')),
                    "age": entry.get('age', 'N/A'),
                    "mac_address": entry.get('mac', entry.get('mac_address', 'N/A')).replace(':', '').replace('-',
                                                                                                              '').replace(
                        '.', ''),
                    "type": entry.get('type', 'ARPA'),
                    "interface": entry.get('interface', 'N/A')})
        else:
            raw_arp_out = net_connect.send_command("show arp", use_textfsm=False, expect_string=r"#")
            if "No ARP entries found" in raw_arp_out or not raw_arp_out.strip() or \
                    ("Total ARP Entries" in raw_arp_out and "0" in raw_arp_out.split("Total ARP Entries")[-1]):
                return []

            header_skipped_arp = False
            for line in raw_arp_out.splitlines():
                line_s = line.strip()
                if not line_s or line_s.lower().startswith("ip address") or \
                        line_s.lower().startswith("total arp") or line_s.lower().startswith("arp entries"):
                    header_skipped_arp = True;
                    continue
                if not header_skipped_arp and not re.match(r"^\s*[\d\.]+", line_s): continue

                match_arp = re.match(r"^\s*([\d\.]+)\s+([0-9a-f\.\:\-]+)\s+(vlan\d+|\S+)\s+.*", line_s, re.IGNORECASE)
                if match_arp:
                    ip, mac, intf_arp = match_arp.groups()
                    arp_table.append({
                        "protocol": "Internet", "address": ip, "age": "N/A",
                        "mac_address": mac.replace(':', '').replace('-', '').replace('.', ''), "type": "ARPA",
                        "interface": intf_arp
                    })
        return arp_table
    except Exception as e:
        print(f"  Erreur critique dans get_arp_table pour {net_connect.host}: {e}");
        traceback.print_exc()
        return []


def check_security_features(net_connect, running_config):
    security_audit = {}
    virtual_prefixes = ("vlan", "loopback", "lag")  # Utilisé pour le check des ports inutilisés

    # --- I. AAA & Authentification & Accès Management ---
    if "aaa authentication port-access" in running_config:
        security_audit["aaa_port_access_configured"] = {"status": True, "level": "good",
                                                        "details": "AAA pour l'accès port (dot1x/mac-auth) semble configuré. Vérifier les détails de la configuration."}
    else:
        security_audit["aaa_port_access_configured"] = {"status": False, "level": "warning",
                                                        "details": "AAA pour l'accès port (dot1x/mac-auth) non détecté. Recommandé pour sécuriser l'accès au réseau."}

    local_user_passwords_encrypted = True  # Par défaut
    if re.search(r"user\s+\S+\s+password\s+plaintext", running_config):
        local_user_passwords_encrypted = False

    if local_user_passwords_encrypted and "password" in running_config:  # Si 'password' est trouvé mais pas 'plaintext'
        security_audit["local_user_password_encryption"] = {"status": "Chiffré (haché)", "level": "good",
                                                            "details": "Les mots de passe des utilisateurs locaux semblent être stockés chiffrés (hachés)."}
    elif not local_user_passwords_encrypted:
        security_audit["local_user_password_encryption"] = {"status": "Plaintext détecté", "level": "bad",
                                                            "details": "Au moins un mot de passe utilisateur local est stocké en clair (plaintext). Utiliser 'password ciphertext <hash>' ou 'password sha256 <hash>'."}
    else:
        security_audit["local_user_password_encryption"] = {
            "status": "Aucun utilisateur local avec mot de passe trouvé ou format inconnu", "level": "warning",
            "details": "Vérifier manuellement le stockage des mots de passe des utilisateurs locaux."}

    # Politique de complexité des mots de passe
    # Note: Aruba OS-CX gère cela via 'password-policy' ou des paramètres individuels.
    # 'show security password-quality' ou 'show password-policy'
    min_len_str, complexity_str = "Non configuré", "Non configurée"
    min_len_level, complexity_level = "bad", "bad"

    if "password minimum-length" in running_config:
        match_pass_len = re.search(r"password minimum-length\s+(\d+)", running_config)
        min_len = int(match_pass_len.group(1)) if match_pass_len else 0
        min_len_str = f"Longueur min: {min_len}"
        min_len_level = "good" if min_len >= 12 else "warning" if min_len >= 8 else "bad"
    security_audit["password_min_length"] = {"status": min_len_str, "level": min_len_level,
                                             "details": f"{min_len_str}. Recommandé: >=12."}

    # Pour la complexité, OS-CX a 'password complexity [level]' ou 'password-policy <name>' -> 'character-class-check'
    # Ceci est une vérification simplifiée.
    if "password complexity" in running_config or "character-class-check" in running_config:
        complexity_str = "Activée (détails à vérifier)"
        complexity_level = "good"
    security_audit["password_complexity"] = {"status": complexity_str, "level": complexity_level,
                                             "details": f"Politique de complexité des mots de passe: {complexity_str}. Vérifier les exigences spécifiques."}

    # Enable password (équivalent à 'enable secret' sur Cisco)
    if "enable password" in running_config:  # OS-CX peut utiliser 'enable password [ciphertext|plaintext] ...'
        if "plaintext" in running_config.split("enable password")[1].splitlines()[0]:
            security_audit["enable_password_aruba"] = {"status": "Plaintext", "level": "bad",
                                                       "details": "Le mot de passe 'enable' est stocké en clair. Utiliser une version chiffrée."}
        else:
            security_audit["enable_password_aruba"] = {"status": "Chiffré", "level": "good",
                                                       "details": "Le mot de passe 'enable' est stocké chiffré."}
    else:
        security_audit["enable_password_aruba"] = {"status": "Non configuré", "level": "bad",
                                                   "details": "Aucun mot de passe 'enable' configuré."}

    # --- II. Sécurité des Lignes d'Accès (Console) ---
    # OS-CX utilise 'line console' puis les paramètres dessous
    console_config_text = ""
    console_match = re.search(r"line console\s*\n(.*?)(?=line|interface|vlan|router|exit|$)", running_config,
                              re.DOTALL | re.MULTILINE)
    if console_match:
        console_config_text = console_match.group(1)

    if "password " in console_config_text or "login local" in console_config_text or "login group" in console_config_text:
        security_audit["console_auth"] = {"status": True, "level": "good",
                                          "details": "Ligne console protégée par méthode de login."}
    else:
        security_audit["console_auth"] = {"status": False, "level": "bad",
                                          "details": "Ligne console non protégée par mot de passe."}

    exec_timeout_con_match = re.search(r"session-timeout\s+(\d+)", console_config_text)  # OS-CX utilise session-timeout
    if exec_timeout_con_match:
        minutes_con = int(exec_timeout_con_match.group(1))
        if minutes_con > 0 and minutes_con <= 15:  # Timeout raisonnable
            security_audit["console_session_timeout"] = {"status": f"{minutes_con} minutes", "level": "good",
                                                         "details": "Timeout de session configuré sur console."}
        elif minutes_con == 0:
            security_audit["console_session_timeout"] = {"status": "Désactivé (0)", "level": "bad",
                                                         "details": "Timeout de session console désactivé. Risque."}
        else:  # Trop long
            security_audit["console_session_timeout"] = {"status": f"{minutes_con} minutes", "level": "warning",
                                                         "details": "Timeout de session console élevé. Recommandé : 5-15 min."}
    else:
        security_audit["console_session_timeout"] = {"status": "Non configuré (défaut)", "level": "warning",
                                                     "details": "Aucun timeout de session sur console. Recommandé : 5-15 min."}

    # --- III. Sécurité des Services de Management ---
    # SSH
    try:
        # Utiliser "show ssh server all-vrfs" pour une vue complète
        ssh_status_raw = net_connect.send_command("show ssh server all-vrfs", expect_string=r"#")
        # print(f"DEBUG: show ssh server all-vrfs output:\n{ssh_status_raw}") # Pour débogage

        ssh_server_is_enabled = False  # On va le déterminer par la présence de la sortie
        ssh_v2_is_primary = False
        ssh_v1_is_active = True  # Supposer actif par défaut si non explicitement désactivé

        if "SSH server configuration on VRF" in ssh_status_raw:  # Indique que le service est configuré
            ssh_server_is_enabled = True  # Si la section existe, le serveur est au moins configuré globalement

        # Chercher la version SSH globale ou par VRF
        # L'image montre "SSH Version : 2.0"
        match_ssh_version = re.search(r"SSH Version\s*:\s*([\d\.]+)", ssh_status_raw)
        if match_ssh_version and match_ssh_version.group(1) == "2.0":
            ssh_v2_is_primary = True

        # Sur OS-CX, SSHv1 est désactivé si 'no ssh server v1 enable' est dans la config
        # ou si 'show ssh server (all-vrfs)' indique explicitement que v1 n'est pas utilisé.
        # La sortie que vous avez fournie ne mentionne pas SSHv1, ce qui est bon signe.
        # Netmiko se connecte en SSH, donc le service est actif.
        # Nous allons vérifier si v1 est explicitement mentionné comme actif ou si v2 n'est pas la seule version.

        # Si la sortie de 'show ssh server all-vrfs' ne contient PAS de référence à SSHv1 actif
        # et que la version est bien 2.0, on considère que c'est bon.
        # Une recherche négative de "SSHv1" ou "Version : 1" serait un indicateur.
        # La sortie que vous avez montrée n'indique PAS de SSHv1.

        # Si on s'est connecté en SSH, le serveur est actif.
        # La question est de savoir si v1 est aussi actif.
        # Aruba OS-CX a tendance à être v2 par défaut et v1 doit être explicitement activé (ou désactivé).
        # Si 'no ssh server v1 enable' est dans la config, c'est une preuve de désactivation.

        v1_disabled_by_config = "no ssh server v1 enable" in running_config
        v1_explicitly_enabled_in_config = "ssh server v1 enable" in running_config  # Moins probable

        if not ssh_server_is_enabled and "ssh" in str(net_connect.device_type).lower():
            # Ce cas est pour si 'show ssh server all-vrfs' ne retourne rien de concluant
            # mais qu'on est connecté en SSH.
            security_audit["ssh_status"] = {"status": "Activé (connexion SSH active)", "level": "warning",
                                            "details": "Serveur SSH fonctionnel, mais détails de version/v1 via 'show ssh server all-vrfs' non clairs. Vérifier manuellement."}
        elif ssh_v2_is_primary and (v1_disabled_by_config or not v1_explicitly_enabled_in_config):
            # Si la version rapportée est 2.0 et que v1 n'est pas explicitement activé dans la config
            # (ou mieux, est explicitement désactivé), on considère que c'est bon.
            security_audit["ssh_status"] = {"status": "SSHv2 Only (probable)", "level": "good",
                                            "details": "Serveur SSH activé, version 2.0 détectée. SSHv1 semble désactivé."}
        elif ssh_v2_is_primary and v1_explicitly_enabled_in_config:
            security_audit["ssh_status"] = {"status": "SSHv2 avec SSHv1 activé", "level": "bad",
                                            "details": "SSHv2 est utilisé, mais SSHv1 est explicitement activé dans la configuration. Désactiver SSHv1."}
        else:  # Cas où SSH est actif mais on n'est pas sûr pour v1 et v2 n'est pas clairement la seule.
            security_audit["ssh_status"] = {"status": "SSH Activé (statut v1 incertain)", "level": "warning",
                                            "details": "Serveur SSH activé, mais impossible de confirmer la désactivation de SSHv1. Assurer 'no ssh server v1 enable'."}

    except Exception as e_ssh:
        print(f"  Erreur durant la vérification SSH pour {net_connect.host}: {e_ssh}")
        security_audit["ssh_status"] = {"status": "Activé (connexion SSH active)", "level": "warning",
                                        "details": "Serveur SSH fonctionnel (connexion établie), mais la commande 'show ssh server all-vrfs' a échoué. Version et statut de SSHv1 inconnus."}

    # Telnet (OS-CX : 'telnet-server enable' ou 'no telnet-server enable')
    if "no telnet-server enable" in running_config or "telnet-server" not in running_config:  # Désactivé par défaut ou explicitement
        security_audit["telnet_server"] = {"status": "Disabled", "level": "good",
                                           "details": "Serveur Telnet désactivé."}
    elif "telnet-server enable" in running_config:
        security_audit["telnet_server"] = {"status": "Enabled", "level": "bad",
                                           "details": "Serveur Telnet activé. Protocole non chiffré, à désactiver."}
    else:  # Cas ambigu
        security_audit["telnet_server"] = {"status": "Statut incertain (non explicitement activé/désactivé)",
                                           "level": "warning",
                                           "details": "Vérifier manuellement le statut du serveur Telnet."}

    # HTTP/HTTPS servers
    https_enabled_in_config = False
    if "https-server" in running_config:
        https_block_match = re.search(r"https-server\s*\n(.*?)(?=^\S|\Z)", running_config, re.DOTALL | re.MULTILINE)
        if https_block_match and "enable" in https_block_match.group(1):
            https_enabled_in_config = True

    if https_enabled_in_config:
        security_audit["https_server"] = {"status": "Enabled", "level": "good",
                                          "details": "Serveur HTTPS (web management) activé."}
    else:
        security_audit["https_server"] = {"status": "Disabled or Not Found", "level": "good",
                                          "details": "Serveur HTTPS non détecté comme activé."}

    http_enabled_in_config = False
    if "http-server" in running_config:  # OS-CX a rarement http-server si https est là
        http_block_match = re.search(r"http-server\s*\n(.*?)(?=^\S|\Z)", running_config, re.DOTALL | re.MULTILINE)
        if http_block_match and "enable" in http_block_match.group(1):
            http_enabled_in_config = True

    if http_enabled_in_config:
        security_audit["http_server"] = {"status": "Enabled", "level": "bad",
                                         "details": "Serveur HTTP (non sécurisé) activé. À désactiver."}
    else:
        security_audit["http_server"] = {"status": "Disabled or Not Found", "level": "good",
                                         "details": "Serveur HTTP non détecté comme activé."}

    # Banners
    banners_set = []
    if "banner motd" in running_config: banners_set.append("MOTD")
    if "banner exec" in running_config: banners_set.append("Exec")  # Moins courant sur OS-CX
    if "banner login" in running_config: banners_set.append("Login")  # Moins courant sur OS-CX
    if banners_set:
        security_audit["banners_configured"] = {"status": f"Configurées: {', '.join(banners_set)}", "level": "good",
                                                "details": "Bannières d'avertissement configurées."}
    else:
        security_audit["banners_configured"] = {"status": "Aucune", "level": "warning",
                                                "details": "Aucune bannière d'avertissement configurée. Recommandé pour des raisons légales."}

    # --- IV. Renforcement Général & Services ---
    # Source Routing (souvent pas une commande explicite 'no ip source-route' sur OS-CX, désactivé par défaut)
    security_audit["ip_source_route_aruba"] = {"status": "Likely Disabled (default)", "level": "good",
                                               "details": "Le routage par la source est généralement désactivé par défaut sur OS-CX."}

    # LLDP/CDP (CDP est rare sur Aruba, LLDP est standard)
    if "no lldp enable" in running_config:  # Si LLDP est explicitement désactivé globalement
        security_audit["lldp_global_status"] = {"status": "Globally Disabled", "level": "good",
                                                "details": "LLDP est désactivé globalement."}
    elif "lldp enable" in running_config or "lldp" in running_config:  # Activé par défaut ou explicitement
        security_audit["lldp_global_status"] = {"status": "Globally Enabled", "level": "warning",
                                                "details": "LLDP est activé globalement. Filtrer sur les interfaces non-confiance via 'no lldp tx-enable/rx-enable'."}
    else:  # Ni l'un ni l'autre, probablement activé par défaut
        security_audit["lldp_global_status"] = {"status": "Potentially Enabled (default)", "level": "warning",
                                                "details": "LLDP est probablement activé par défaut. À vérifier."}

    # --- V. Logging & Monitoring ---
    log_level_details = "N/A"
    log_level_match = re.search(r"logging severity\s+(\S+)", running_config)
    if log_level_match: log_level_details = f"Sévérité globale: {log_level_match.group(1)}"

    if "logging syslog host" in running_config or "logging host" in running_config:  # 'logging host <IP>' sur OS-CX
        security_audit["remote_logging_aruba"] = {"status": True, "level": "good",
                                                  "details": f"Logging distant (syslog) configuré. {log_level_details}"}
        if "logging source-interface" in running_config and "loopback" in running_config:  # Bonne pratique
            src_int_match = re.search(r"logging source-interface\s+(\S+)", running_config)
            security_audit["logging_source_int_aruba"] = {
                "status": src_int_match.group(1) if src_int_match else "Configurée", "level": "good",
                "details": "Interface source pour syslog spécifiée (Loopback est bien)."}
        else:
            security_audit["logging_source_int_aruba"] = {"status": False, "level": "warning",
                                                          "details": "Interface source pour syslog non spécifiée ou pas une loopback."}
    else:
        security_audit["remote_logging_aruba"] = {"status": False, "level": "bad",
                                                  "details": f"Logging distant (syslog) NON configuré. {log_level_details}"}
        security_audit["logging_source_int_aruba"] = {"status": "N/A", "level": "bad",
                                                      "details": "Syslog non configuré."}

    # NTP
    ntp_servers_aruba = [line for line in running_config.splitlines() if line.strip().startswith("ntp server")]
    num_ntp_aruba = len(ntp_servers_aruba)
    ntp_sync_level_aruba, ntp_sync_details_aruba, ntp_sync_status_aruba = "warning", "Statut synchro NTP inconnu.", "Error"
    try:
        ntp_status_aruba_raw = net_connect.send_command("show ntp status", expect_string=r"#")
        if "Clock is synchronized" in ntp_status_aruba_raw:
            stratum_match = re.search(r"stratum\s+(\d+)", ntp_status_aruba_raw)
            ntp_sync_status_aruba, ntp_sync_details_aruba, ntp_sync_level_aruba = True, f"NTP synchronisé. Stratum: {stratum_match.group(1) if stratum_match else 'N/A'}.", "good"
        else:
            ntp_sync_status_aruba, ntp_sync_details_aruba = False, "NTP non synchronisé."
            ntp_sync_level_aruba = "bad"  # Non synchro est toujours mauvais
    except:
        ntp_sync_details_aruba = "Statut synchro inconnu ('show ntp status' échoué)."
        ntp_sync_level_aruba = "warning" if num_ntp_aruba > 0 else "bad"
    security_audit["ntp_synchronization_aruba"] = {"status": ntp_sync_status_aruba, "level": ntp_sync_level_aruba,
                                                   "details": ntp_sync_details_aruba}

    if num_ntp_aruba >= 2:
        security_audit["ntp_redundancy_aruba"] = {"status": f"{num_ntp_aruba} serveurs", "level": "good",
                                                  "details": "Redondance NTP OK."}
    elif num_ntp_aruba == 1:
        security_audit["ntp_redundancy_aruba"] = {"status": "1 serveur", "level": "warning",
                                                  "details": "Un seul serveur NTP. Recommandé: >=2."}
    else:
        security_audit["ntp_redundancy_aruba"] = {"status": "0 serveur", "level": "bad",
                                                  "details": "NTP non configuré. Temps non fiable."}

    # SNMP
    if "snmp-server community public" in running_config or "snmp-server community private" in running_config:
        security_audit["snmp_default_communities_aruba"] = {"status": True, "level": "bad",
                                                            "details": "Communautés SNMP par défaut (public/private) utilisées. Risque majeur."}
    elif "snmp-server community" in running_config:
        security_audit["snmp_default_communities_aruba"] = {"status": False, "level": "good",
                                                            "details": "Communautés SNMP personnalisées. Vérifier les ACLs associées."}
    else:
        security_audit["snmp_default_communities_aruba"] = {"status": "N/A", "level": "good",
                                                            "details": "Pas de communautés SNMP v1/v2c configurées."}

    # SNMPv3 (OS-CX utilise 'snmp-server vrf <vrf> user <user> auth ...')
    snmpv3_user_found = "snmp-server user " in running_config or "snmp-server vrf " in running_config and " user " in running_config  # Recherche simplifiée

    if snmpv3_user_found:
        security_audit["snmp_v3_aruba"] = {"status": True, "level": "good",
                                           "details": "SNMPv3 semble configuré (utilisateurs/groupes détectés)."}
    elif "snmp-server community" in running_config:  # Si pas de v3 mais v1/v2c
        security_audit["snmp_v3_aruba"] = {"status": "v1/v2c only", "level": "bad",
                                           "details": "SNMPv1/v2c utilisé (communautés en clair). Préférer SNMPv3."}
    else:
        security_audit["snmp_v3_aruba"] = {"status": "N/A", "level": "good",
                                           "details": "SNMP (v1/v2c/v3) ne semble pas configuré."}

    # --- VI. Sécurité Couche 2 (Indications) ---
    # OS-CX: 'spanning-tree port <port> bpdu-protection'
    if "bpdu-protection" in running_config:  # Recherche globale
        security_audit["bpdu_protection_aruba"] = {"status": True, "level": "good",
                                                   "details": "BPDU Protection (Guard) semble être configurée sur certaines interfaces."}
    else:
        security_audit["bpdu_protection_aruba"] = {"status": False, "level": "warning",
                                                   "details": "BPDU Protection non détectée. Recommandé sur les ports d'accès."}

    # OS-CX: 'dhcp-snooping enable' et 'dhcp-snooping vlan <vlan-id> enable'
    if "dhcp-snooping enable" in running_config and "dhcp-snooping vlan" in running_config:
        security_audit["dhcp_snooping_aruba"] = {"status": True, "level": "good",
                                                 "details": "DHCP Snooping semble activé globalement et pour des VLANs."}
    elif "dhcp-snooping" in running_config:  # Partiellement configuré
        security_audit["dhcp_snooping_aruba"] = {"status": "Partiel", "level": "warning",
                                                 "details": "DHCP Snooping partiellement configuré. Vérifier l'activation globale ET par VLAN."}
    else:
        security_audit["dhcp_snooping_aruba"] = {"status": False, "level": "bad",
                                                 "details": "DHCP Snooping non activé. Requis pour prévenir les serveurs DHCP pirates."}

    # Port Security (MAC Authentication / dot1x) déjà couvert par 'aaa_port_access_configured'
    # Storm control (appelé 'rate-limit' dans OS-CX pour broadcast/multicast/unknown-unicast)
    if "rate-limit " in running_config and (
            "broadcast " in running_config or "multicast " in running_config or "unknown-unicast " in running_config):
        security_audit["rate_limiting_bcast_mcast"] = {"status": True, "level": "good",
                                                       "details": "Limitation de débit (storm control) pour broadcast/multicast/unknown-unicast semble configurée."}
    else:
        security_audit["rate_limiting_bcast_mcast"] = {"status": False, "level": "warning",
                                                       "details": "Limitation de débit (storm control) non détectée. Utile contre les tempêtes de trafic."}

    return security_audit


def load_inventory(filepath="inventory.csv"):
    inventory = []
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
                if len(header) < 3:
                    print(
                        f"Erreur: En-tête inventaire '{filepath}' doit avoir au moins 3 colonnes (hostname, group, device_type).")
                    return None
            except StopIteration:
                print(f"Avertissement: Fichier inventaire '{filepath}' vide.");
                return []
            for row in reader:
                if len(row) >= 3 and row[0].strip():
                    inventory.append(
                        {"host": row[0].strip(), "group": row[1].strip(), "device_type": row[2].strip().lower()})
                elif row and any(field.strip() for field in row):
                    print(f"Avertissement: Ligne inventaire mal formatée: {row}")
        return inventory
    except FileNotFoundError:
        print(f"Erreur: Fichier inventaire '{filepath}' non trouvé.");
        return None
    except Exception as e:
        print(f"Erreur lecture '{filepath}': {e}");
        traceback.print_exc();
        return None


def load_passwords(filepath="passwords.csv"):
    passwords = {}
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            try:
                next(reader)
            except StopIteration:
                print(f"Avertissement: Fichier passwords '{filepath}' vide.");
                return {}
            for row in reader:
                if len(row) >= 3 and row[0].strip():
                    enable_pass = row[3].strip() if len(row) > 3 and row[3].strip() else None
                    passwords[row[0].strip()] = {"username": row[1].strip(), "password": row[2].strip(),
                                                 "enable_password": enable_pass}
                elif row and any(field.strip() for field in row):
                    print(f"Avertissement: Ligne passwords mal formatée: {row}")
        return passwords
    except FileNotFoundError:
        print(f"Erreur: Fichier passwords '{filepath}' non trouvé.");
        return None
    except Exception as e:
        print(f"Erreur lecture '{filepath}': {e}");
        traceback.print_exc();
        return None


def generate_excel_report(all_data, excel_filepath):
    if not all_data: print(f"Aucune donnée pour rapport Excel : {excel_filepath}."); return
    wb = openpyxl.Workbook();
    wb.remove(wb.active)
    ws_info = wb.create_sheet("Infos Générales")
    headers_info = ["Hostname", "IP Address", "Modèle", "Version OS", "Uptime", "Numéro de Série"]
    ws_info.append(headers_info);
    apply_header_style(ws_info)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection':
            ws_info.append(
                [dev_data.get('attempted_host', 'N/A'), dev_data.get('attempted_host', 'N/A'), "ERREUR CONNEXION",
                 dev_data.get('error_message', 'N/A'), "N/A", "N/A"])
            for i in range(3, 5): ws_info.cell(row=ws_info.max_row, column=i).fill = RED_FILL
        else:
            info = dev_data.get("general_info", {})
            ws_info.append([info.get("hostname", dev_data.get("host", "N/A")), dev_data.get("host", "N/A"),
                            info.get("model", "N/A"), info.get("ios_version", "N/A"), info.get("uptime", "N/A"),
                            info.get("serial_number", "N/A")])
    auto_fit_columns(ws_info)
    ws_interfaces = wb.create_sheet("Interfaces")
    headers_interfaces = ["Hostname", "Nom Interface", "Type", "Description", "IP Address", "Statut Lien",
                          "Statut Protocole", "VLAN (Access)", "Duplex", "Vitesse"]
    ws_interfaces.append(headers_interfaces);
    apply_header_style(ws_interfaces)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for iface in dev_data.get("interfaces", []):
            ws_interfaces.append(
                [hostname, iface.get("name"), iface.get("type"), iface.get("description"), iface.get("ip_address"),
                 iface.get("status_link"), iface.get("status_protocol"), iface.get("vlan"), iface.get("duplex"),
                 iface.get("speed")])
            lr, pr = ws_interfaces.cell(row=ws_interfaces.max_row, column=6), ws_interfaces.cell(
                row=ws_interfaces.max_row, column=7)
            sl, sp = str(iface.get("status_link", "")).lower(), str(iface.get("status_protocol", "")).lower()
            if sl == "up" or "connected" in sl:
                lr.fill = GREEN_FILL
            elif ("admin" in sl and "down" in sl) or sl == "disabled" or "admin-down" in sl:
                lr.fill = ORANGE_FILL
            elif sl == "down" or "notconnect" in sl or "err-disabled" in sl or "error-disabled" in sl or "down (no transceiver)" in sl:
                lr.fill = RED_FILL
            if sp == "up":
                pr.fill = GREEN_FILL
            elif sp == "down":
                pr.fill = RED_FILL
    auto_fit_columns(ws_interfaces)
    ws_vlans = wb.create_sheet("VLANs")
    headers_vlans = ["Hostname", "ID VLAN", "Nom VLAN", "Statut", "Ports Assignés"]
    ws_vlans.append(headers_vlans);
    apply_header_style(ws_vlans)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for vlan in dev_data.get("vlans", []):
            ws_vlans.append([hostname, vlan.get("id"), vlan.get("name"), vlan.get("status"), vlan.get("ports")])
            sc = ws_vlans.cell(row=ws_vlans.max_row, column=4)
            if str(vlan.get("status", "")).lower() == "up" or str(vlan.get("status", "")).lower() == "active":
                sc.fill = GREEN_FILL
            else:
                sc.fill = ORANGE_FILL
    auto_fit_columns(ws_vlans)
    ws_arp = wb.create_sheet("Table ARP")
    headers_arp = ["Hostname", "Protocole", "Adresse IP", "Âge (min)", "Adresse MAC", "Type", "Interface"]
    ws_arp.append(headers_arp);
    apply_header_style(ws_arp)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for entry in dev_data.get("arp_table", []):
            ws_arp.append(
                [hostname, entry.get("protocol"), entry.get("address"), entry.get("age"), entry.get("mac_address"),
                 entry.get("type"), entry.get("interface")])
    auto_fit_columns(ws_arp)
    ws_security = wb.create_sheet("Audit Sécurité")
    headers_security = ["Hostname", "Point de Contrôle", "Statut/Valeur", "Niveau", "Détails/Recommandation"]
    ws_security.append(headers_security);
    apply_header_style(ws_security)
    for dev_data in all_data:
        if dev_data.get('status') == 'error_connection': continue
        hostname = dev_data.get("general_info", {}).get("hostname", dev_data.get("host", "N/A"))
        for check_name, check_data in dev_data.get("security_audit", {}).items():
            ws_security.append(
                [hostname, check_name.replace("_", " ").title(), str(check_data.get("status")), check_data.get("level"),
                 check_data.get("details")])
            lc = ws_security.cell(row=ws_security.max_row, column=4)
            set_cell_status_color(lc, check_data.get("level"))
    auto_fit_columns(ws_security)
    try:
        wb.save(excel_filepath);
        print(f"\n[+] Rapport Excel Aruba généré : {excel_filepath}")
    except Exception as e:
        print(f"\n[-] Erreur sauvegarde Excel Aruba: {e}")


def perform_aruba_audit(aruba_devices_inventory, global_passwords_map, output_directory):
    all_devices_data = []
    for device_entry in aruba_devices_inventory:
        host, group = device_entry["host"], device_entry["group"]
        creds = global_passwords_map.get(group)
        print(f"\n[INFO Aruba] Traitement de {host} (groupe: {group})...")

        if not creds:
            print(f"  [ERREUR Aruba] Identifiants non trouvés pour groupe '{group}'. {host} ignoré.")
            all_devices_data.append({"attempted_host": host, "status": "error_connection",
                                     "error_message": f"Identifiants non trouvés pour groupe {group}"})
            continue

        device_params = {
            'device_type': 'aruba_aoscx_ssh',
            'host': host, 'username': creds['username'], 'password': creds['password'],
            'secret': creds.get('enable_password'),
            'global_delay_factor': 2, 'timeout': 45, 'session_timeout': 120
        }
        current_device_data = {"host": host}
        try:
            with ConnectHandler(**device_params) as net_connect:
                actual_host = net_connect.host
                actual_prompt = (
                    net_connect.base_prompt[:-1] if net_connect.base_prompt and net_connect.base_prompt.endswith(
                        ('#', '>')) else actual_host)
                print(f"  [OK Aruba] Connecté à {actual_host} ({actual_prompt}).")

                current_device_data["general_info"] = get_aruba_device_info(net_connect)
                current_device_data["general_info"]["ip_address_queried"] = host

                running_config = net_connect.send_command("show running-config", read_timeout=240, expect_string=r"#")
                if not running_config: running_config = ""

                current_device_data["interfaces"] = get_aruba_interfaces(net_connect)
                current_device_data["vlans"] = get_vlans(net_connect)
                current_device_data["arp_table"] = get_arp_table(net_connect)
                current_device_data["security_audit"] = check_security_features(net_connect, running_config)
                all_devices_data.append(current_device_data)

        except (NetmikoTimeoutException, SSHException) as e:
            print(f"  [ERREUR Aruba] Connexion à {host} (Timeout/SSH): {e}")
            all_devices_data.append({"attempted_host": host, "status": "error_connection", "error_message": str(e)})
        except NetmikoAuthenticationException as e:
            print(f"  [ERREUR Aruba] Authentification sur {host}: {e}")
            all_devices_data.append(
                {"attempted_host": host, "status": "error_connection", "error_message": f"Échec authentification: {e}"})
        except Exception as e:
            if "Unsupported 'device_type'" in str(e):
                print(
                    f"  [ERREUR Aruba] Unsupported 'device_type' pour {host}: {e}. Vérifiez le 'device_type' ('aruba_aoscx_ssh') ou la version de Netmiko.")
                all_devices_data.append({"attempted_host": host, "status": "error_connection",
                                         "error_message": f"Unsupported 'device_type': {e}"})
            else:
                print(f"  [ERREUR Aruba] Inattendue avec {host}: {e}");
                traceback.print_exc()
                all_devices_data.append(
                    {"attempted_host": host, "status": "error_connection", "error_message": f"Erreur inattendue: {e}"})

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = os.path.join(output_directory, f"audit_aruba_data_{timestamp}.json")
    try:
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(all_devices_data, f, indent=4, ensure_ascii=False)
        print(f"\n[+] Données JSON Aruba sauvegardées : {json_filename}")
    except Exception as e:
        print(f"\n[-] Erreur sauvegarde JSON Aruba: {e}")

    excel_report_path = os.path.join(output_directory, f"audit_aruba_report_{timestamp}.xlsx")
    generate_excel_report(all_devices_data, excel_report_path)

    print(f"\n[+] Audit Aruba terminé pour {len(aruba_devices_inventory)} équipement(s).")


def main_aruba():
    inventory_file, password_file, output_directory = "inventory.csv", "passwords.csv", "audit_reports"
    if not os.path.exists(output_directory):
        try:
            os.makedirs(output_directory)
        except OSError as e:
            print(f"Erreur création répertoire '{output_directory}': {e}");
            return

    full_inventory = load_inventory(inventory_file)
    passwords_map = load_passwords(password_file)

    if full_inventory is None or passwords_map is None:
        print("Arrêt (Aruba): Erreurs critiques chargement fichiers entrée.");
        return

    aruba_devices = [device for device in full_inventory if device.get("device_type") == "aruba_os-cx"]

    if not aruba_devices:
        print("Aucun équipement Aruba OS-CX trouvé dans l'inventaire pour audit autonome.")
        return

    perform_aruba_audit(aruba_devices, passwords_map, output_directory)


if __name__ == "__main__":
    main_aruba()
